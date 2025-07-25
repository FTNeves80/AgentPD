import openpyxl

inv_file = openpyxl.load_workbook('inventory.xlsx')
product_list = inv_file['Sheet1']

products_per_supplier = {}

for product_row in range(2,product_list.max_row + 1):
    supplier_name = product_list.cell(product_row,4).value
    
    #calculation number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        print(f"New supplier found: {supplier_name}")
        products_per_supplier[supplier_name] = 1

print(f"Supplier_name: {products_per_supplier}")
